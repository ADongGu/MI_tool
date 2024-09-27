import openpyxl
import sys
from openpyxl import Workbook
from openpyxl.styles import Font
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QTableWidget, QTableWidgetItem, QVBoxLayout, QTextEdit,QMessageBox,QRadioButton

from PyQt5.QtWidgets import QMainWindow, QApplication, QPushButton, QVBoxLayout, QWidget, QFileDialog,QGraphicsView, QGraphicsScene

class OrderItem:
    #序号	货物编号	规格（长宽高）MM	  单位	数量	单价	金额	采购订单
    def __init__(self,
                 serial_number: int,
                 product_code: str ,
                 dimensions:str,
                 unit:int , quantity:int,
                 unit_price:float, amount:float,
                 purchase_order:str):
        self.serial_number = serial_number
        self.product_code = product_code
        self.dimensions = dimensions
        self.unit = unit
        self.quantity = quantity
        self.unit_price = unit_price
        self.amount = amount
        self.purchase_order = purchase_order

    def __repr__(self):
        return (f"OrderItem("
                f"serial_number={self.serial_number}, "
                f"product_code={self.product_code}, "
                f"dimensions={self.dimensions}, "
                f"unit={self.unit}, "
                f"quantity={self.quantity}, "
                f"unit_price={self.unit_price}, "
                f"amount={self.amount}, "
                f"purchase_order={self.purchase_order})")
class SongHuo_Sheet:
    def __init__(self):
        self.id = ""
        self.m_no = ""
        self.m_date = ""
        self.m_ddnumber = ""
        self.orderItems = []
class DeliveryItem:
    #序号	订单号	送货日期	送货单号	名称	模号	规格	单位	数量	单价	金 额 (RMB)
    def __init__(self,
                 serial_number: int,
                 m_ddnumber: str,
                 delivery_date: str,
                 delivery_note_number: str,
                 name: str,
                 model: str,
                 specification: str,
                 unit: str,
                 quantity: int,
                 unit_price: float,
                 amount_rmb: float):
        self.serial_number = serial_number
        self.m_ddnumber = m_ddnumber
        self.delivery_date = delivery_date
        self.delivery_note_number = delivery_note_number
        self.name = name
        self.model = model
        self.specification = specification
        self.unit = unit
        self.quantity = quantity
        self.unit_price = unit_price
        self.amount_rmb = amount_rmb

    def __repr__(self):
        return (f"DeliveryItem("
                f"serial_number={self.serial_number}, "
                f"m_ddnumber='{self.m_ddnumber}', "
                f"delivery_date='{self.delivery_date}', "
                f"delivery_note_number='{self.delivery_note_number}', "
                f"name='{self.name}', "
                f"model='{self.model}', "
                f"specification='{self.specification}', "
                f"unit='{self.unit}', "
                f"quantity={self.quantity}, "
                f"unit_price={self.unit_price}, "
                f"amount_rmb={self.amount_rmb})")
def convert_date_format(date_str):
    # 去除字符串中的非日期字符，只保留日期部分
    date_part = date_str.split('：')[1]  # 去除'出货日期：'部分
    date_part = date_part.replace('年', '.')  # 将'年'替换为'-'
    date_part = date_part.replace('月', '.')  # 将'月'替换为'-'
    date_part = date_part.replace('日', '')  # 去除'日'字符

    # 返回格式化后的日期字符串
    return date_part

G_SongHuo_sheets = []
G_DuiZhang_sheets = []

def get_songhuo_sheet_old(path):
    type = 0

    # 创建工作簿和活动工作表
    wb_d = Workbook()
    ws_d = wb_d.active

    for i in range(0, 8):
        ws_d.append([])
    headers = ["序号", "订单号", "送货日期", "送货单号", "名称", "模号", "规格", "单位", "数量", "单价", "金额 (RMB)"]
    ws_d.append(headers)

    # 打开一个Excel文件
    wb = openpyxl.load_workbook(path, data_only=True)


    # 遍历每个工作表
    number = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]  # 获取当前工作表对象

        tmp_sheet = SongHuo_Sheet()
        tmp_sheet.id = sheet_name

        try:
            s =str(sheet.cell(row=2,column=8).value)
            if len(s) <= 4:
                raise Exception("这是一个主动抛出的异常")
            else:
                s1 = str(sheet.cell(row=2,column=8).value)
                ss = str(sheet.cell(row=2,column=8).value)[0]
                tmp_sheet.m_no = str(sheet.cell(row=2,column=8).value)[3:]
        except Exception as e:
            try:
                tmp_sheet.m_no = str(sheet.cell(row=2, column=9).value)[3:]
            except Exception as e:
                tmp_sheet.m_no = "无填写，请检查"

        try:
            tmp_sheet.m_date = convert_date_format(str(sheet.cell(row=4,column=7).value))
        except Exception as e:
            try:
                tmp_sheet.m_date = convert_date_format(str(sheet.cell(row=3, column=7).value))
                type = 1
            except Exception as e:
                tmp_sheet.m_date = "无填写，请检查"

        try:
            tmp_sheet.m_ddnumber = str(sheet.cell(row=3,column=4).value).split('：')[1]
        except Exception as e:
            tmp_sheet.m_ddnumber = "无填写，请检查"

        for row in sheet.iter_rows(min_row=6-int(type), max_row=13, max_col=sheet.max_column):  # 假设所有行的列数相同
            # 创建一个列表来保存当前行的数据
            row_data = [cell.value for cell in row]
            if row_data[2] == "以下空白" or row_data[3] == None:
                continue
            tmp = OrderItem(row_data[0], row_data[1],row_data[2],row_data[3],row_data[4],row_data[5],row_data[6],row_data[7])
            tmp_sheet.orderItems.append((tmp))

            # tmp_del = DeliveryItem(number+1,tmp_sheet.m_no,tmp_sheet.m_date,tmp_sheet.m_ddnumber,"木箱",
            #                        row_data[1], row_data[2], row_data[3], row_data[4], row_data[5], row_data[6]
            #                        )
            ws_d.append([number+1,tmp_sheet.m_ddnumber,tmp_sheet.m_date,tmp_sheet.m_no,"木箱",
                                   row_data[1], row_data[2], row_data[3], row_data[4], row_data[5], row_data[6]])
            number = number + 1
        wb_d.save('科航类型的对账单.xlsx')

        G_SongHuo_sheets.append(tmp_sheet)

    # 关闭工作簿
    wb.close()

def get_songhuo_sheet(path):
    type = 0

    # 创建工作簿和活动工作表
    wb_d = Workbook()
    ws_d = wb_d.active

    # for i in range(0, 4):
    #     ws_d.append([])
    ws_d.append(["东莞市长安精固木箱厂"])
    ws_d.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    ws_d.append(["2024年月份7对账单"])
    ws_d.merge_cells(start_row=2, start_column=1, end_row=2, end_column=8)
    ws_d.append(["客 户："])
    ws_d.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)
    ws_d.append(["联系人："])
    ws_d.merge_cells(start_row=4, start_column=1, end_row=4, end_column=3)
    ws_d.merge_cells(start_row=4, start_column=6, end_row=4, end_column=8)
    ws_d['F4'] = '日期：2024-11-14'
    headers = ["序号", "交货日期", "送货单号",  "模号", "规格及型号",  "数量", "单价(RMB)", "金额(RMB)", "采购人"]
    ws_d.append(headers)

    # 打开一个Excel文件
    wb = openpyxl.load_workbook(path, data_only=True)


    # 遍历每个工作表
    number = 0
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]  # 获取当前工作表对象

        tmp_sheet = SongHuo_Sheet()
        tmp_sheet.id = sheet_name

        try:
            s =str(sheet.cell(row=2,column=8).value)
            if len(s) <= 4:
                raise Exception("这是一个主动抛出的异常")
            else:
                s1 = str(sheet.cell(row=2,column=8).value)
                ss = str(sheet.cell(row=2,column=8).value)[0]
                tmp_sheet.m_no = str(sheet.cell(row=2,column=8).value)[3:]
        except Exception as e:
            try:
                tmp_sheet.m_no = str(sheet.cell(row=2, column=9).value)[3:]
            except Exception as e:
                tmp_sheet.m_no = "无填写，请检查"


        try:
            tmp_sheet.m_date = convert_date_format(str(sheet.cell(row=4,column=7).value))
        except Exception as e:
            try:
                tmp_sheet.m_date = convert_date_format(str(sheet.cell(row=3,column=7).value))
                type = 1
            except Exception as e:
                tmp_sheet.m_date = "无填写，请检查"



        try:
            tmp_sheet.m_ddnumber = str(sheet.cell(row=3,column=4).value).split('：')[1]
        except Exception as e:
            tmp_sheet.m_ddnumber = "无填写，请检查"

        for row in sheet.iter_rows(min_row=6-int(type), max_row=13, max_col=sheet.max_column):  # 假设所有行的列数相同
            # 创建一个列表来保存当前行的数据
            row_data = [cell.value for cell in row]
            if row_data[2] == "以下空白" or row_data[3] == None:
                continue
            tmp = OrderItem(row_data[0], row_data[1],row_data[2],row_data[3],row_data[4],row_data[5],row_data[6],row_data[7])
            tmp_sheet.orderItems.append((tmp))

            # tmp_del = DeliveryItem(number+1,tmp_sheet.m_no,tmp_sheet.m_date,tmp_sheet.m_ddnumber,"木箱",
            #                        row_data[1], row_data[2], row_data[3], row_data[4], row_data[5], row_data[6]
            #                        )

            ws_d.append([number+1,tmp_sheet.m_date,tmp_sheet.m_no,
                                   row_data[1], row_data[2],  row_data[4], row_data[5], row_data[6], str(sheet.cell(row=3,column=1).value)])
            number = number + 1
        wb_d.save('自己简洁的对账单.xlsx')

        G_SongHuo_sheets.append(tmp_sheet)

    # 关闭工作簿
    wb.close()

def get_duizhang_sheet(path):
    # 打开一个Excel文件
    wb = openpyxl.load_workbook(path, data_only=True)
    # 遍历每个工作表
    for sheet_name in wb.sheetnames:
        print(f"Processing sheet: {sheet_name}")
        sheet = wb[sheet_name]  # 获取当前工作表对象

        for row in sheet.iter_rows(min_row=10,  max_col=sheet.max_column):  # 假设所有行的列数相同
            row_data = [cell.value for cell in row]
            if row_data[0] == '合计金额（大写）：':
                break
            # print(row_data[0])
            tmp = DeliveryItem(row_data[0], row_data[1],row_data[2],row_data[3],
                               row_data[4],row_data[5],row_data[6],row_data[7],
                               row_data[8],row_data[9],row_data[10])
            # print(tmp.__repr__())
            # print()
            G_DuiZhang_sheets.append(tmp)
            # G_DuiZhang_sheets.orderItems.append((tmp))


    # 关闭工作簿
    wb.close()

class App(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()
        self.name1 = ""
        self.name2 =""
        self.file_type = "生成科航类型的对账单"

        self.setAcceptDrops(True)  # 设置窗口接受拖拽事件

    def initUI(self):
        # 创建单选框
        self.radio1 = QRadioButton('生成科航类型的对账单')
        self.radio1.setChecked(True)
        self.radio2 = QRadioButton('生成自己简洁的对账单')

        # 信号和槽函数连接
        self.radio1.toggled.connect(self.onRadioToggled)
        self.radio2.toggled.connect(self.onRadioToggled)


        # 创建两个按钮，用于打开文件选择对话框
        self.btn_open_file1 = QPushButton('选择送货单', self)
        self.btn_open_file1.clicked.connect(self.open_file1)

        # 创建布局并添加按钮和表格
        layout = QVBoxLayout()
        layout.addWidget(self.btn_open_file1)
        layout.addWidget(self.radio1)
        layout.addWidget(self.radio2)

        # 设置布局
        self.setLayout(layout)

        # 设置窗口标题和初始大小
        self.setWindowTitle('生成对账单')
        self.setGeometry(300, 300, 600, 400)

    def onRadioToggled(self):
        # 获取触发事件的单选框
        radio = self.sender()
        if radio.isChecked():
            print(radio.text() + ' 被选中')
            self.file_type = radio.text()

    def open_file1(self):
        # 使用QFileDialog.getOpenFileName弹出文件选择对话框
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog  # 可以取消使用系统原生文件选择对话框
        fileName, _ = QFileDialog.getOpenFileName(self, f"{'选择送货单'} 文件", "", "All Files (*);;Text Files (*.txt)",
                                                  options=options)
        print(fileName)

        if fileName:
            global  G_SongHuo_sheets
            G_SongHuo_sheets = []
            # 将选择的文件路径添加到表格中
            # self.text_edit.append(fileName)
            self.name1 = fileName
            if(self.file_type == "生成自己简洁的对账单"):
                get_songhuo_sheet(fileName)
            else:
                get_songhuo_sheet_old(fileName)
            QMessageBox.information(self, '信息', '已经生成好啦')

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():  # 检查是否有URLs数据
            for url in event.mimeData().urls():
                if url.isLocalFile() and self.is_image_file(url.toLocalFile()):  # 检查是否是本地视频文件
                    event.accept()
                    return
        event.ignore()

    def dropEvent(self, event):
        for url in event.mimeData().urls():
            if url.isLocalFile() and self.is_image_file(url.toLocalFile()):
                file_path = url.toLocalFile()
                print(file_path)
                global G_SongHuo_sheets
                G_SongHuo_sheets = []
                # 将选择的文件路径添加到表格中
                # self.text_edit.append(fileName)
                self.name1 = file_path
                if (self.file_type == "生成自己简洁的对账单"):
                    get_songhuo_sheet(file_path)
                else:
                    get_songhuo_sheet_old(file_path)
                QMessageBox.information(self, '信息', '已经生成好啦')
    def is_image_file(self, file_path):
        import os
        valid_extensions = ['.xlsx']
        if os.path.isfile(file_path) and os.path.splitext(file_path)[1] in valid_extensions:
            return True
        return False


app = QApplication(sys.argv)
if __name__ == '__main__':

    ex = App()
    ex.show()
    sys.exit(app.exec_())



